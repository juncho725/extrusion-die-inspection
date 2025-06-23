import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from PyQt5.QtWidgets import QPushButton, QFileDialog, QMessageBox
from datetime import datetime

class ExcelReportGenerator:
    def __init__(self, parent=None):
        self.parent = parent
        self._setup_template_path()

    def _setup_template_path(self):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        project_root = os.path.dirname(os.path.dirname(current_dir))
        self.template_path = os.path.join(project_root, "src", "resources", "templates", "inspection_report_template.xlsx")
    
    def create_report_button(self):
        """Create a button for generating Excel reports"""
        button = QPushButton('Generate Excel Report')
        button.setFixedHeight(40)
        button.clicked.connect(self.generate_excel_report)
        return button

    def generate_excel_report(self):
        """Generate Excel report from the template and current data"""
        try:
            print("===== Excel Report Generation Started =====")
            def format_decimal_3(value):
                """Function to display values with exactly 3 decimal places"""
                return float(f"{float(value):.3f}")
            # Define thin border style (used only when border settings are needed)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Maintain basic validation code...
            if not self.parent or not hasattr(self.parent, 'current_results_dir'):
                QMessageBox.warning(None, "Warning", "App instance not found.")
                return
            
            # Important: Error occurred here! Define csv_path variable
            csv_path = os.path.join(self.parent.current_results_dir, "vertical_object_distances.csv")
            print(f"Debug: CSV file path = {csv_path}")
            if not os.path.exists(csv_path):
                QMessageBox.warning(None, "Warning", "Distance data file not found.")
                return
            print(f"Debug: CSV file exists: {csv_path}")
            
            # Load the CSV data
            df = pd.read_csv(csv_path,dtype=str, encoding='utf-8-sig')
            print(f"Debug: CSV data loaded, rows: {len(df)}, columns: {len(df.columns)}")
            df.iloc[1:, 1:] = df.iloc[1:, 1:].astype(float)
            # Prepare dataframe for Excel export
            df_excel = df.copy(deep=True)
            
            # Process product columns
            product_columns = [col for col in df_excel.columns if col.startswith('Product_')]
            product_columns.sort(key=lambda x: int(x.split('_')[1]))
            print(f"Debug: Product column list: {product_columns}")
            
            # Maintain number conversion logic...
            
            # Load template
            wb = openpyxl.load_workbook(self.template_path)
            ws = wb.active
            print(f"Debug: Template workbook loaded, active sheet: {ws.title}")

            # ── [Initialize background color] Remove all cell Fill in A1~Z70 range ──
            no_fill = PatternFill(fill_type=None)

            for row in ws['A1':'Z70']:        # Inspection range
                for cell in row:
                    if cell.fill and cell.fill.fill_type not in (None, 'none'):
                        cell.fill = no_fill

            # Set save path - Add this code
            today = datetime.now().strftime("%Y%m%d")
            default_filename = f"inspection_report_{today}.xlsx"

            if not hasattr(self.parent, 'current_session_dir'):
                print("Debug: self.parent does not have current_session_dir attribute. Using default path")
                save_dir = os.path.dirname(self.parent.current_results_dir)  # Use parent folder of results directory
            else:
                save_dir = self.parent.current_session_dir
                
            save_path, _ = QFileDialog.getSaveFileName(
                self.parent, 
                "Save Report", 
                os.path.join(save_dir, default_filename),
                "Excel Files (*.xlsx)"
            )

            if not save_path:
                print("Debug: User cancelled file save.")
                return
                
            print(f"Debug: Save path = {save_path}")
            # Set production information
            # Set creation date (B4)
            if hasattr(self.parent, 'production_date_input'):
                prod_date = self.parent.production_date_input.date()
                ws['C4'] = prod_date.toString("yyyy-MM-dd")
                print(f"Debug: Creation date set: {prod_date.toString('yyyy-MM-dd')}")
                
            # Set coating date (B6)
            if hasattr(self.parent, 'coating_date_input'):
                coat_date = self.parent.coating_date_input.date()
                ws['C6'] = coat_date.toString("yyyy-MM-dd")
                print(f"Debug: Coating date set: {coat_date.toString('yyyy-MM-dd')}")
                
            # Set die number (D7)
            if hasattr(self.parent, 'dice_number_input'):
                dice_number = self.parent.dice_number_input.text()
                ws['D7'] = dice_number
                print(f"Debug: Die number set: {dice_number}")
            
            # Set engraving numbers (G14, I14, etc)
            # Get engraving numbers (row with Distance_Category = 'Product Name')
            engraving_row = df[df['Distance_Category'] == 'Product Name']
            print(f"Debug: Engraving number row exists: {not engraving_row.empty}")
            
            # Fill engraving numbers in row 14
            col_offset = 0
            for i, product_col in enumerate(product_columns):
                if not engraving_row.empty and product_col in engraving_row.columns:
                    engraving_no = engraving_row.iloc[0][product_col]
                    if pd.notna(engraving_no):
                        # Each engraving number takes 2 cells horizontally
                        col_letter = openpyxl.utils.get_column_letter(7 + col_offset)  # g is column 7

                        cell = ws[f"{col_letter}14"]
                        cell.value = str(engraving_no)     
                        cell.number_format = '@'              

                        col_offset += 2
            
            # Get session settings
            target_distances = {}
            tolerances_plus = {}
            tolerances_minus = {}
            
            if hasattr(self.parent, 'session_settings'):
                if 'target_distances' in self.parent.session_settings:
                    target_distances = self.parent.session_settings.get("target_distances", {})
                    print(f"Debug: target_distances = {target_distances}")
                if 'tolerances_plus' in self.parent.session_settings:
                    tolerances_plus = self.parent.session_settings.get("tolerances_plus", {})
                    print(f"Debug: tolerances_plus = {tolerances_plus}")
                if 'tolerances_minus' in self.parent.session_settings:
                    tolerances_minus = self.parent.session_settings.get("tolerances_minus", {})
                    print(f"Debug: tolerances_minus = {tolerances_minus}")
            
            # Set default values
            default_values = {
                'A': {'target': 15.37, 'plus': 0.010, 'minus': 0.010},
                'B': {'target': 1.23, 'plus': 0.010, 'minus': 0.010},
                'E': {'target': 0.25, 'plus': 0.010, 'minus': 0.010},
                'F': {'target': 0.25, 'plus': 0.010, 'minus': 0.010},
                'G': {'target': 0.24, 'plus': 0.010, 'minus': 0.010},
            }
            
            # Upper section row mapping
            upper_section_mapping = {
                'A00': 15,        # 01 A (rows 15-16)
                'B00': 17,        # 02 B (rows 17-18)
                'E01': 19,        # 03 E (rows 19-20)
                'F02': 21,        # 04 F (rows 21-22)
                'G03': 23,        # 05 G (rows 23-24)
                'G04': 25,        # 06 G (rows 25-26)
                'G05': 27,        # 07 G (rows 27-28)
                'G06': 29,        # 08 G (rows 29-30)
                'G07': 31,        # 09 G (rows 31-32)
                'G08': 33,        # 10 G (rows 33-34)
                'G09': 35,        # 11 G (rows 35-36)
                'G10': 37,        # 12 G (rows 37-38)
                'G11': 39,        # 13 G (rows 39-40)
                'G12': 41,        # 14 G (rows 41-42)
                'G13': 43,        # 15 G (rows 43-44)
                'G14': 45,        # 16 G (rows 45-46)
                'G15': 47,        # 17 G (rows 47-48)
                'F16': 49,        # 18 F (rows 49-50)
                'E17': 51,        # 19 E (rows 51-52)
            }
            
            # Lower section row mapping (newly added)
            lower_section_mapping = {
                'A': 55,   # Lower section A row
                'B': 57,   # Lower section B row
                'E': 59,   # Lower section E row
                'F': 61,   # Lower section F row
                'G': 63,   # Lower section G row
            }
            
            # Map simplified categories to upper section as well
            simplified_upper_mapping = {
                'A': 15,  # A same position as A00
                'B': 17,  # B same position as B00
                'E': 19,  # E same position as E01
                'F': 21,  # F same position as F02
                'G': 23,  # G same position as G03
            }
            upper_section_mapping.update(simplified_upper_mapping)
            
            # Generate tolerance row mapping
            tolerance_row_mapping = {}
            for cat, row in upper_section_mapping.items():
                tolerance_row_mapping[cat] = (row, row+1)
            
            # 1. Process upper section
            print("\n===== Upper Section Processing Started =====")
            for csv_category, excel_row in upper_section_mapping.items():
                print(f"\nDebug: Processing category {csv_category}, Excel row: {excel_row}")
                
                # Extract base category
                base_category = csv_category[0] if len(csv_category) > 0 else csv_category
                
                # Get target dimension and tolerance values
                # Get values from GUI settings
                if base_category in target_distances:
                    target = target_distances[base_category]
                    print(f"Debug: Got target dimension from GUI settings for {base_category}: {target}")
                else:
                    target = default_values[base_category]['target']
                    print(f"Debug: Got target dimension from default values for {base_category}: {target}")
                    
                if base_category in tolerances_plus:
                    plus_tol = tolerances_plus[base_category]
                    print(f"Debug: Got + tolerance from GUI settings for {base_category}: {plus_tol}")
                else:
                    plus_tol = default_values[base_category]['plus']
                    print(f"Debug: Got + tolerance from default values for {base_category}: {plus_tol}")
                    
                if base_category in tolerances_minus:
                    minus_tol = tolerances_minus[base_category]
                    print(f"Debug: Got - tolerance from GUI settings for {base_category}: {minus_tol}")
                else:
                    minus_tol = default_values[base_category]['minus']
                    print(f"Debug: Got - tolerance from default values for {base_category}: {minus_tol}")
                
                # In the upper section processing part, change this:
                try:
                    # Set target dimension (E column)
                    ws.cell(row=excel_row, column=5).value = float(target)  # E column is 5th column
                    ws.cell(row=excel_row, column=5).number_format = '0.000'
                    print(f"Debug: Upper section - target dimension set complete - cell E{excel_row}, value: {target}")
                    
                    # Tolerance row numbers
                    tol_row_plus, tol_row_minus = tolerance_row_mapping[csv_category]
                    
                    # Set tolerance (C column) - positive sign
                    ws.cell(row=tol_row_plus, column=3).value = "+"  # Set sign only in C column
                    ws.cell(row=tol_row_plus, column=3).alignment = Alignment(horizontal='right')
                    
                    # Set tolerance (D column) - positive value
                    ws.cell(row=tol_row_plus, column=4).value = format_decimal_3(plus_tol)  # Set value only in D column
                    ws.cell(row=tol_row_plus, column=4).number_format = '0.000'
                    print(f"Debug: Upper section - + tolerance set complete - cell D{tol_row_plus}, value: {format_decimal_3(plus_tol)}")
                    
                    # Set tolerance (C column) - negative sign
                    ws.cell(row=tol_row_minus, column=3).value = "-"  # Set sign only in C column
                    ws.cell(row=tol_row_minus, column=3).alignment = Alignment(horizontal='right')
                    
                    # Set tolerance (D column) - negative value
                    ws.cell(row=tol_row_minus, column=4).value = format_decimal_3(minus_tol)  # Set value only in D column (as positive)
                    ws.cell(row=tol_row_minus, column=4).number_format = '0.000'
                    print(f"Debug: Upper section - - tolerance set complete - cell D{tol_row_minus}, value: {format_decimal_3(minus_tol)}")
                except Exception as e:
                    print(f"Debug: Upper section value setting failed: {str(e)}")
                # Process measurement value data
                category_rows = df_excel[df_excel['Distance_Category'] == csv_category]
                
                # If no exact match, search for similar categories
                if category_rows.empty:
                    # Find other categories starting with single character (A, B, E, F, G)
                    if csv_category[0] == csv_category:
                        alt_categories = [c for c in df_excel['Distance_Category'].unique() 
                                            if c.startswith(csv_category)]
                        if alt_categories:
                            category_rows = df_excel[df_excel['Distance_Category'] == alt_categories[0]]
                
                # Input measurement values
                if not category_rows.empty:
                    for j, product_col in enumerate(product_columns):
                        if product_col in category_rows.columns and pd.notna(category_rows.iloc[0][product_col]):
                            value = category_rows.iloc[0][product_col]
                            
                            # Calculate column position (each measurement value takes 2 columns)
                            col = 7 + j * 2  # G column is 7th
                            col_letter = openpyxl.utils.get_column_letter(col)
                            
                            # Set value (no merge operation)
                            ws.cell(row=excel_row, column=col).value = value
                            
                            # Set number format
                            if isinstance(value, (int, float)):
                                ws.cell(row=excel_row, column=col).number_format = '0.000'
                            
                            print(f"Debug: Measurement value set complete - cell {col_letter}{excel_row}, value: {value}")
            
            # Save workbook
            print(f"\nDebug: Workbook save started: {save_path}")
            wb.save(save_path)
            print(f"Debug: Workbook save complete: {save_path}")
            
            QMessageBox.information(self.parent, "Success", f"Report generated successfully.\nSave path: {save_path}")
            print("===== Excel Report Generation Complete =====")
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            print(f"Error occurred: {str(e)}\n{error_details}")
            QMessageBox.critical(self.parent, "Error", f"Error during report generation: {str(e)}\n\n{error_details}")